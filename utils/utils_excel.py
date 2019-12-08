# -*- coding: utf-8 -*-

import logging

from utils.utils_type import unicode2utf8
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelInvalidException(Exception):
    def __init__(self, errs):
        self.errs = errs

    def __str__(self):
        return 'Excel contains invalid data'


class ExcelErr(object):

    def __init__(self, row, col, err, top):
        self.row = row + 1
        self.col = str(colnum_to_name(col)).upper()
        self.err = err
        self.sort = 1 if top else 100

    def __str__(self):
        if self.row == self.col == -1:
            return unicode2utf8(self.err)
        return unicode2utf8(u'行%d，列%s，错误：%s' % (self.row, self.col, self.err))


class ExcelErrsMixin(object):

    def __init__(self):
        self.err_list = list()

    def clean_err(self):
        self.err_list[:] = []

    def add_err(self, row, col, err, top=False):
        err = ExcelErr(row, col, err, top)
        logger.warn(err)
        self.err_list.append(err)

    def retrieve_err(self):
        return sorted(self.err_list, key=lambda each: each.sort)


class ExcelFileStreamLoadMixin(object):

    def __init__(self, stream):
        from io import BytesIO
        # wb = load_workbook(filename=BytesIO(input_excel.read()))
        self.book = load_workbook(filename=BytesIO(stream.read()))
        self.sheet = self.book.active
        return


class ExcelFileLocalLoadMixin(object):

    def __init__(self, fpath):
        pass


def colname_to_num(colname):
    if type(colname) is not str:
        return colname
    col = 0
    power = 1
    for i in xrange(len(colname) - 1, -1, -1):
        ch = colname[i]
        col += (ord(ch) - ord('A') + 1) * power
        power *= 26
    return col - 1


def colnum_to_name(colnum):
    if type(colnum) != int:
        return colnum
    if colnum == -1:
        return -1
    if colnum > 25:
        ch1 = chr(colnum % 26 + 65)
        ch2 = chr(colnum / 26 + 64)
        return ch2 + ch1
    else:
        return chr(colnum % 26 + 65)


